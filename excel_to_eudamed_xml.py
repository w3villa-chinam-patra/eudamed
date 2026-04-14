#!/usr/bin/env python3
"""Generate EUDAMED XML from Excel and validate against the XSD schema.

This version builds the XML tree directly from Excel data and does not depend on
an XML template file.
"""

from __future__ import annotations

import argparse
import os
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path

import openpyxl
from lxml import etree

from eudamed_constants import HEADERS, SHEET_PREFIXES, XML_NS, XML_TAGS


NS = XML_NS

XML_NS_MAP = {
    "message": NS["message"],
    "service": NS["service"],
    "device": NS["device"],
    "e": NS["e"],
    "budi": NS["budi"],
    "udidi": NS["udidi"],
    "commondevice": NS["commondevice"],
    "lngs": NS["lngs"],
    "mktinfo": NS["mktinfo"],
    "lnks": NS["lnks"],
    "xsi": "http://www.w3.org/2001/XMLSchema-instance",
}

# XSD-safe defaults used only when required fields are blank.
DEFAULT_CONVERSATION_ID = "00000000-0000-0000-0000-000000000000"
DEFAULT_CORRELATION_ID = "00000000-0000-0000-0000-000000000001"
DEFAULT_MESSAGE_ID = "00000000-0000-0000-0000-000000000002"
DEFAULT_RECIPIENT_NODE_ACTOR_CODE = "IT-MF-000000001"
DEFAULT_SENDER_NODE_ACTOR_CODE = "EUDAMED_MDR"
DEFAULT_MF_ACTOR_CODE = "IT-MF-000000001"
DEFAULT_UDI_STATUS_CODE = "ON_THE_MARKET"
DEFAULT_MDN_CODE = "A0000"
DEFAULT_REFERENCE_NUMBER = "REF-001"

XML_FIXED = {
    "state_registered": "REGISTERED",
    "version": "1",
    "device_type": "DEVICE",
    "service_id": "DEVICE",
    "service_operation": "GET",
    "response_code": "SUCCESS",
}


RISK_CLASS_MAP = {
    "class 1": "CLASS_I",
    "class i": "CLASS_I",
    "class 2a": "CLASS_IIA",
    "class iia": "CLASS_IIA",
    "class 2b": "CLASS_IIB",
    "class iib": "CLASS_IIB",
    "class 3": "CLASS_III",
    "class iii": "CLASS_III",
}

STATUS_MAP = {
    "on the eu market": "ON_THE_MARKET",
    "on the market": "ON_THE_MARKET",
    "no longer placed on the eu market": "NO_LONGER_PLACED_ON_THE_MARKET",
    "no longer placed on the market": "NO_LONGER_PLACED_ON_THE_MARKET",
    "not intended for the eu market": "NOT_INTENDED_FOR_EU_MARKET",
}

LANGUAGE_MAP = {
    "all languages": "ANY",
    "any": "ANY",
    "english": "EN",
    "french": "FR",
    "german": "DE",
    "dutch": "NL",
    "spanish": "ES",
    "bulgarian": "BG",
    "czech": "CS",
    "danish": "DA",
    "estonian": "ET",
    "greek": "EL",
    "irish": "GA",
    "croatian": "HR",
    "italian": "IT",
    "latvian": "LV",
    "lithuanian": "LT",
    "hungarian": "HU",
    "maltese": "MT",
    "polish": "PL",
    "portuguese": "PT",
    "romanian": "RO",
    "slovak": "SK",
    "slovenian": "SL",
    "finnish": "FI",
    "swedish": "SV",
    "icelandic": "IS",
    "norwegian": "NO",
    "turkish": "TR",
}

# Declarative Excel -> XML boolean mappings.
# Format: (xml_tag_key, source_sheet, excel_header, default_value)
BASIC_UDI_BOOL_MAPPINGS = [
    ("budi.animal_tissues_cells", "device", HEADERS["animal_tissues"], "false"),
    ("budi.human_tissues_cells", "device", HEADERS["human_tissues"], "false"),
    ("budi.human_product_check", "device", HEADERS["human_blood_product"], "false"),
    ("budi.medicinal_product_check", "device", HEADERS["medicinal_product"], "false"),
    ("commondevice.active", "basic", HEADERS["active"], "false"),
    ("commondevice.administering_medicine", "basic", HEADERS["administering_medicine"], "false"),
    ("commondevice.implantable", "basic", HEADERS["implantable"], "false"),
    ("commondevice.measuring_function", "basic", HEADERS["measuring_function"], "false"),
    ("commondevice.reusable", "basic", HEADERS["reusable"], "false"),
]

UDIDI_BOOL_MAPPINGS = [
    ("udidi.sterile", "chars", HEADERS["sterile"], "false"),
    ("udidi.sterilization", "chars", HEADERS["sterilization"], "false"),
    ("udidi.latex", "chars", HEADERS["latex"], "false"),
    ("udidi.reprocessed", "device", HEADERS["reprocessed"], "false"),
]


@dataclass
class SheetData:
    headers: list[str]
    values: list[object]

    def get(self, header: str, occurrence: int = 1) -> str:
        hits = [i for i, h in enumerate(self.headers) if (h or "").strip().lower() == header.strip().lower()]
        if len(hits) < occurrence:
            return ""
        value = self.values[hits[occurrence - 1]]
        if value is None:
            return ""
        return str(value).strip()


@dataclass
class SchemaConstraints:
    risk_classes: set[str]
    device_statuses: set[str]
    issuing_entities: set[str]
    certificate_types: set[str]
    basic_required_fields: set[str]
    udidi_required_fields: set[str]


class ExcelToEudamedXML:
    def __init__(self, workspace_root: Path):
        self.workspace_root = workspace_root

    def _resolve(self, path: str) -> Path:
        resolved = Path(path)
        return resolved if resolved.is_absolute() else (self.workspace_root / resolved)

    @staticmethod
    def _to_bool(value: str | object) -> str:
        raw = str(value or "").strip().lower()
        if raw in {"y", "yes", "true", "1"}:
            return "true"
        if raw in {"n", "no", "false", "0"}:
            return "false"
        return ""

    @staticmethod
    def _normalize_risk_class(value: str) -> str:
        raw = (value or "").strip().lower()
        return RISK_CLASS_MAP.get(raw, (value or "").strip().replace(" ", "_").upper())

    @staticmethod
    def _normalize_status(value: str) -> str:
        raw = (value or "").strip().lower()
        return STATUS_MAP.get(raw, (value or "").strip().replace(" ", "_").upper())

    @staticmethod
    def _normalize_language(value: str, default: str = "ANY") -> str:
        raw = (value or "").strip()
        if not raw:
            return default
        upper_raw = raw.upper()
        if len(upper_raw) in {2, 3}:
            return upper_raw
        return LANGUAGE_MAP.get(raw.lower(), default)

    @staticmethod
    def _normalize_certificate_type(value: str) -> str:
        raw = (value or "").strip()
        return raw.upper().replace("-", "_").replace(" ", "_")

    @staticmethod
    def _normalize_nb_actor_code(value: str) -> str:
        raw = (value or "").strip()
        digits = "".join(character for character in raw if character.isdigit())
        if len(digits) == 4:
            return digits
        return raw

    @staticmethod
    def _sheet_data(ws: openpyxl.worksheet.worksheet.Worksheet) -> SheetData:
        headers = [str(ws.cell(1, c).value).strip() if ws.cell(1, c).value is not None else "" for c in range(1, ws.max_column + 1)]
        values = [ws.cell(2, c).value for c in range(1, ws.max_column + 1)] if ws.max_row >= 2 else []
        return SheetData(headers=headers, values=values)

    def _sheet_by_prefix(self, wb: openpyxl.Workbook, prefix: str) -> openpyxl.worksheet.worksheet.Worksheet:
        for name in wb.sheetnames:
            if name.lower().startswith(prefix.lower()):
                return wb[name]
        raise ValueError(f"Required sheet starting with '{prefix}' not found")

    def _optional_sheet_by_prefix(self, wb: openpyxl.Workbook, prefix: str) -> openpyxl.worksheet.worksheet.Worksheet | None:
        for name in wb.sheetnames:
            if name.lower().startswith(prefix.lower()):
                return wb[name]
        return None

    @staticmethod
    def _ns_tag(prefix: str, local_name: str) -> str:
        return f"{{{NS[prefix]}}}{local_name}"

    @staticmethod
    def _tag_key_parts(tag_key: str) -> tuple[str, str]:
        return XML_TAGS[tag_key]

    def _ns_tag_key(self, tag_key: str) -> str:
        prefix, local_name = self._tag_key_parts(tag_key)
        return self._ns_tag(prefix, local_name)

    def _append(self, parent: etree._Element, prefix: str, local_name: str, text: str | None = None) -> etree._Element:
        element = etree.SubElement(parent, self._ns_tag(prefix, local_name))
        if text is not None:
            element.text = text
        return element

    def _append_tag(self, parent: etree._Element, tag_key: str, text: str | None = None) -> etree._Element:
        prefix, local_name = self._tag_key_parts(tag_key)
        return self._append(parent, prefix, local_name, text)

    def _append_bool(self, parent: etree._Element, prefix: str, local_name: str, raw_value: str, default: str = "false") -> etree._Element:
        value = self._to_bool(raw_value)
        if value == "":
            value = default
        return self._append(parent, prefix, local_name, value)

    def _append_bool_tag(self, parent: etree._Element, tag_key: str, raw_value: str, default: str = "false") -> etree._Element:
        prefix, local_name = self._tag_key_parts(tag_key)
        return self._append_bool(parent, prefix, local_name, raw_value, default)

    def _append_required_text(self, parent: etree._Element, prefix: str, local_name: str, value: str, error_message: str) -> etree._Element:
        if not value:
            raise ValueError(error_message)
        return self._append(parent, prefix, local_name, value)

    def _append_required_text_tag(self, parent: etree._Element, tag_key: str, value: str, error_message: str) -> etree._Element:
        if not value:
            raise ValueError(error_message)
        return self._append_tag(parent, tag_key, value)

    def _append_text_or_comment(
        self,
        parent: etree._Element,
        prefix: str,
        local_name: str,
        value: str,
        comment_if_blank: str,
    ) -> etree._Element:
        text = (value or "").strip()
        if text:
            return self._append(parent, prefix, local_name, text)
        parent.append(etree.Comment(comment_if_blank))
        return self._append(parent, prefix, local_name, "")

    def _append_text_with_default_comment(
        self,
        parent: etree._Element,
        prefix: str,
        local_name: str,
        value: str,
        default_value: str,
        comment_if_defaulted: str,
    ) -> etree._Element:
        text = (value or "").strip()
        if text:
            return self._append(parent, prefix, local_name, text)
        parent.append(etree.Comment(comment_if_defaulted))
        return self._append(parent, prefix, local_name, default_value)

    @staticmethod
    def _sheet_value(sheets: dict[str, SheetData], source_sheet: str, header: str) -> str:
        sheet = sheets.get(source_sheet)
        if sheet is None:
            return ""
        return sheet.get(header)

    def _append_mapped_booleans(
        self,
        parent: etree._Element,
        sheets: dict[str, SheetData],
        mappings: list[tuple[str, str, str, str]],
    ) -> None:
        for tag_key, source_sheet, excel_header, default_value in mappings:
            raw_value = self._sheet_value(sheets, source_sheet, excel_header)
            self._append_bool_tag(parent, tag_key, raw_value, default=default_value)

    @staticmethod
    def _extract_simple_type_enums(xsd_file: Path, type_name: str) -> set[str]:
        if not xsd_file.exists():
            return set()

        doc = etree.parse(str(xsd_file))
        xs_ns = {"xs": "http://www.w3.org/2001/XMLSchema"}
        values = doc.xpath(
            f"//xs:simpleType[@name='{type_name}']/xs:restriction/xs:enumeration/@value",
            namespaces=xs_ns,
        )
        return {str(value).strip() for value in values if str(value).strip()}

    @staticmethod
    def _type_local_name(type_name: str) -> str:
        if not type_name:
            return ""
        return type_name.split(":", 1)[-1]

    @staticmethod
    def _min_occurs(element: etree._Element) -> int:
        value = element.get("minOccurs")
        if value is None:
            return 1
        try:
            return int(value)
        except ValueError:
            return 1

    def _find_complex_type(self, docs: list[etree._ElementTree], type_name: str) -> etree._Element | None:
        local = self._type_local_name(type_name)
        xs_ns = {"xs": "http://www.w3.org/2001/XMLSchema"}
        for doc in docs:
            found = doc.xpath(f"//xs:complexType[@name='{local}']", namespaces=xs_ns)
            if found:
                return found[0]
        return None

    def _find_group(self, docs: list[etree._ElementTree], group_name: str) -> etree._Element | None:
        local = self._type_local_name(group_name)
        xs_ns = {"xs": "http://www.w3.org/2001/XMLSchema"}
        for doc in docs:
            found = doc.xpath(f"//xs:group[@name='{local}']", namespaces=xs_ns)
            if found:
                return found[0]
        return None

    def _collect_required_from_sequence(self, seq: etree._Element, docs: list[etree._ElementTree], out: set[str]) -> None:
        xs_ns = {"xs": "http://www.w3.org/2001/XMLSchema"}
        for child in seq:
            tag = etree.QName(child.tag).localname
            if tag == "element" and self._min_occurs(child) > 0:
                name = child.get("name")
                if name:
                    out.add(name)
            elif tag == "group" and self._min_occurs(child) > 0:
                ref = child.get("ref")
                if not ref:
                    continue
                group = self._find_group(docs, ref)
                if group is None:
                    continue
                group_seq = group.find("xs:sequence", namespaces=xs_ns)
                if group_seq is not None:
                    self._collect_required_from_sequence(group_seq, docs, out)

    def _collect_required_fields_for_type(self, docs: list[etree._ElementTree], type_name: str, visited: set[str] | None = None) -> set[str]:
        if visited is None:
            visited = set()
        local = self._type_local_name(type_name)
        if local in visited:
            return set()
        visited.add(local)

        required: set[str] = set()
        xs_ns = {"xs": "http://www.w3.org/2001/XMLSchema"}
        complex_type = self._find_complex_type(docs, local)
        if complex_type is None:
            return required

        extension = complex_type.find("xs:complexContent/xs:extension", namespaces=xs_ns)
        if extension is not None:
            base = extension.get("base")
            if base:
                required |= self._collect_required_fields_for_type(docs, base, visited)
            seq = extension.find("xs:sequence", namespaces=xs_ns)
            if seq is not None:
                self._collect_required_from_sequence(seq, docs, required)
            return required

        seq = complex_type.find("xs:sequence", namespaces=xs_ns)
        if seq is not None:
            self._collect_required_from_sequence(seq, docs, required)
        return required

    def _required_field_sets(self, xsd_root: Path) -> tuple[set[str], set[str]]:
        docs = [
            etree.parse(str(xsd_root / "data/Entity/Device/RegulationDevice/BasicUDIType.xsd")),
            etree.parse(str(xsd_root / "data/Entity/Device/RegulationDevice/UDIDIType.xsd")),
            etree.parse(str(xsd_root / "data/Entity/Device/CommonDeviceType.xsd")),
            etree.parse(str(xsd_root / "data/Entity/Entity.xsd")),
            etree.parse(str(xsd_root / "data/Entity/Links/LinkType.xsd")),
        ]

        basic_required = self._collect_required_fields_for_type(docs, "MDRBasicUDIType")
        udidi_required = self._collect_required_fields_for_type(docs, "MDRUDIDIDataType")

        # Expand required nested fields for complex types used in required elements.
        if "identifier" in basic_required:
            basic_required.add("identifier.DICode")
            basic_required.add("identifier.issuingEntityCode")
        if "identifier" in udidi_required:
            udidi_required.add("identifier.DICode")
            udidi_required.add("identifier.issuingEntityCode")
        if "basicUDIIdentifier" in udidi_required:
            udidi_required.add("basicUDIIdentifier.DICode")
            udidi_required.add("basicUDIIdentifier.issuingEntityCode")
        if "status" in udidi_required:
            udidi_required.add("status.code")

        return basic_required, udidi_required

    def _load_schema_constraints(self, xsd_file: Path) -> SchemaConstraints:
        xsd_root = xsd_file.parent.parent
        risk_file = xsd_root / "data/Entity/Common/RiskClassEnum.xsd"
        udidi_file = xsd_root / "data/Entity/Device/RegulationDevice/UDIDIType.xsd"
        link_file = xsd_root / "data/Entity/Links/LinkType.xsd"
        basic_required, udidi_required = self._required_field_sets(xsd_root)

        return SchemaConstraints(
            risk_classes=self._extract_simple_type_enums(risk_file, "RiskClassEnum"),
            device_statuses=self._extract_simple_type_enums(udidi_file, "DeviceStatusEnum"),
            issuing_entities=self._extract_simple_type_enums(udidi_file, "IssuingEntityTypeEnum"),
            certificate_types=self._extract_simple_type_enums(link_file, "GenericCertificateTypeEnum"),
            basic_required_fields=basic_required,
            udidi_required_fields=udidi_required,
        )

    def _build_source_values(self, basic: SheetData, ident: SheetData, chars: SheetData, device: SheetData) -> dict[str, str]:
        status = self._normalize_status(ident.get(HEADERS["udi_di_status"]))
        return {
            "riskClass": self._normalize_risk_class(basic.get(HEADERS["risk_class"])),
            "model": basic.get(HEADERS["device_model"]),
            "modelName": basic.get(HEADERS["device_name"]),
            "identifier.DICode": basic.get("Basic UDI-DI code") or basic.get("Basic UDI code") or ident.get(HEADERS["udi_di_code"]),
            "identifier.issuingEntityCode": basic.get(HEADERS["issuing_entity"]) or ident.get(HEADERS["issuing_entity"]),
            "animalTissuesCells": self._to_bool(device.get(HEADERS["animal_tissues"])) or "false",
            "humanTissuesCells": self._to_bool(device.get(HEADERS["human_tissues"])) or "false",
            "MFActorCode": basic.get(HEADERS["manufacturer_srn"]) or basic.get(HEADERS["mf_actor_code"]),
            "humanProductCheck": self._to_bool(device.get(HEADERS["human_blood_product"])) or "false",
            "medicinalProductCheck": self._to_bool(device.get(HEADERS["medicinal_product"])) or "false",
            "type": XML_FIXED["device_type"],
            "active": self._to_bool(basic.get(HEADERS["active"])) or "false",
            "administeringMedicine": self._to_bool(basic.get(HEADERS["administering_medicine"])) or "false",
            "implantable": self._to_bool(basic.get(HEADERS["implantable"])) or "false",
            "measuringFunction": self._to_bool(basic.get(HEADERS["measuring_function"])) or "false",
            "reusable": self._to_bool(basic.get(HEADERS["reusable"])) or "false",
            "status.code": status,
            "basicUDIIdentifier.DICode": basic.get("Basic UDI-DI code") or basic.get("Basic UDI code") or ident.get(HEADERS["udi_di_code"]),
            "basicUDIIdentifier.issuingEntityCode": basic.get(HEADERS["issuing_entity"]) or ident.get(HEADERS["issuing_entity"]),
            "MDNCodes": ident.get(HEADERS["emdn_code"]),
            "referenceNumber": ident.get(HEADERS["reference_number"]) or ident.get(HEADERS["udi_di_code"]),
            "sterile": self._to_bool(chars.get(HEADERS["sterile"])) or "false",
            "sterilization": self._to_bool(chars.get(HEADERS["sterilization"])) or "false",
            "numberOfReuses": (
                "0"
                if self._to_bool(chars.get(HEADERS["single_use"])) == "true"
                else (str(chars.get(HEADERS["max_reuses"])) if str(chars.get(HEADERS["max_reuses"]) or "").isdigit() else "1")
            ),
            "latex": self._to_bool(chars.get(HEADERS["latex"])) or "false",
            "reprocessed": self._to_bool(device.get(HEADERS["reprocessed"])) or "false",
            # New conditional business rule fields
            "isSystemSPP": self._to_bool(basic.get(HEADERS["is_system_spp"])) or "false",
            "isKit": self._to_bool(basic.get(HEADERS["is_kit"])) or "false",
            "specialDeviceType": basic.get(HEADERS["special_device_type"]) or "",
            "iibImplantableExceptions": self._to_bool(basic.get(HEADERS["iib_exception_flag"])) or "false",
            "iibImplantableDeviceType": basic.get(HEADERS["iib_exception_type"]) or "",
            "memberStatePlacement": ident.get(HEADERS["member_state_placement"]) or "",
            "memberStatesMadeAvailable": ident.get(HEADERS["member_states_available"]) or "",
            "additionalDescription": ident.get(HEADERS["additional_description"]) or "",
        }

    def _validate_input_data(self, basic: SheetData, ident: SheetData, chars: SheetData, device: SheetData, cert: SheetData, constraints: SchemaConstraints) -> None:
        errors: list[str] = []
        warnings: list[str] = []
        
        source_values = self._build_source_values(basic, ident, chars, device)

        risk_raw = basic.get(HEADERS["risk_class"])
        risk_class = self._normalize_risk_class(risk_raw)
        if not risk_class:
            errors.append("'Risk class*' is required.")
        elif constraints.risk_classes and risk_class not in constraints.risk_classes:
            allowed = ", ".join(sorted(constraints.risk_classes))
            errors.append(f"Risk class '{risk_raw}' normalized to '{risk_class}' is not valid per XSD. Allowed: {allowed}")

        if not (basic.get(HEADERS["device_model"]) or basic.get(HEADERS["device_name"])):
            errors.append("Either 'Device Model' or 'Device Name*' must be provided.")

        issuing_entity = ident.get(HEADERS["issuing_entity"])
        if not issuing_entity:
            errors.append("'UDI-DI identification Issuing Entity (GS1/HIBCC/ICCBBA/IFA)' is required.")
        elif constraints.issuing_entities and issuing_entity.strip().upper() not in constraints.issuing_entities:
            allowed = ", ".join(sorted(constraints.issuing_entities))
            errors.append(f"Issuing Entity '{issuing_entity}' is not valid per XSD. Allowed: {allowed}")

        if not ident.get(HEADERS["udi_di_code"]):
            errors.append("'UDI-DI code' is required.")

        status_raw = ident.get(HEADERS["udi_di_status"])
        status = self._normalize_status(status_raw)
        if not status:
            errors.append("'UDI-DI status (...)' is required.")
        elif constraints.device_statuses and status not in constraints.device_statuses:
            allowed = ", ".join(sorted(constraints.device_statuses))
            errors.append(f"UDI-DI status '{status_raw}' normalized to '{status}' is not valid per XSD. Allowed: {allowed}")

        cert_applicable = self._to_bool(cert.get(HEADERS["cert_applicable"]))
        if risk_class == "CLASS_III" and cert_applicable != "true":
            errors.append("Certificate information is required for Class III devices. Set 'Certificate applicable' to Yes.")

        if cert_applicable == "true":
            nb_actor_code = self._normalize_nb_actor_code(cert.get(HEADERS["nb_actor_code"]))
            if not nb_actor_code:
                errors.append("'Notified Body Actor Code (NBActorCode)' is required when certificate is applicable.")
            elif not (len(nb_actor_code) == 4 and nb_actor_code.isdigit()):
                errors.append("NBActorCode must be exactly 4 digits as per XSD pattern [0-9]{4}.")

            cert_type = self._normalize_certificate_type(cert.get(HEADERS["cert_type"]))
            if not cert_type:
                errors.append("'Certificate type' is required when certificate is applicable.")
            elif constraints.certificate_types and cert_type not in constraints.certificate_types:
                allowed = ", ".join(sorted(constraints.certificate_types))
                errors.append(f"Certificate type '{cert_type}' is not valid per XSD. Allowed: {allowed}")

        # Dynamic required-field checks driven by XSD minOccurs.
        basic_required = set(constraints.basic_required_fields)
        udidi_required = set(constraints.udidi_required_fields)

        # Container complex elements are validated through their required child fields.
        basic_required.discard("identifier")
        udidi_required.discard("identifier")
        udidi_required.discard("basicUDIIdentifier")
        udidi_required.discard("status")

        # These are intentionally allowed to be blank so output can contain user prompts.
        basic_required.discard("MFActorCode")
        udidi_required.discard("MDNCodes")
        udidi_required.discard("referenceNumber")

        # Choice in BasicUDIType is model OR modelName; treat as one logical requirement.
        if "model" in basic_required or "modelName" in basic_required:
            if not (source_values.get("model") or source_values.get("modelName")):
                errors.append("XSD requires one of 'model' or 'modelName'; provide Device Model or Device Name*.")
            basic_required.discard("model")
            basic_required.discard("modelName")

        for field in sorted(basic_required):
            value = source_values.get(field, "")
            if value in {"", None}:
                errors.append(f"Missing required field for MDRBasicUDIType (from XSD): {field}")

        for field in sorted(udidi_required):
            value = source_values.get(field, "")
            if value in {"", None}:
                errors.append(f"Missing required field for MDRUDIDIDataType (from XSD): {field}")

        # ===== Additional Conditional Business Rules (BR-UDID) =====
        # BR-UDID-676: Class I → Implantable forced to False
        if risk_class == "CLASS_I":
            implantable = self._to_bool(basic.get(HEADERS["implantable"]))
            if implantable == "true":
                warnings.append("BR-UDID-676: Class I devices cannot be implantable. Auto-correcting to No in generated XML.")

        # BR-UDID-677: Implantable → Reusable surgical instrument forced to False
        implantable = self._to_bool(basic.get(HEADERS["implantable"]))
        if implantable == "true":
            reusable = self._to_bool(basic.get(HEADERS["reusable"]))
            if reusable == "true":
                warnings.append("BR-UDID-677: Implantable devices cannot be reusable surgical instruments. Auto-correcting reusable to No in generated XML.")

        # BR-UDID-024: Max reuses only if NOT single-use
        single_use = self._to_bool(chars.get(HEADERS["single_use"]))
        if single_use == "true":
            max_reuses_raw = chars.get(HEADERS["max_reuses"])
            if max_reuses_raw and str(max_reuses_raw).strip() and int(str(max_reuses_raw).strip() or "0") > 0:
                errors.append("BR-UDID-024: Maximum number of reuses cannot be provided for single-use devices.")

        # BR-UDID-043: Member States required for Class IIa/IIb/III when on market
        if risk_class in {"CLASS_IIA", "CLASS_IIB", "CLASS_III"} and status == "ON_THE_MARKET":
            member_state_placement = ident.get(HEADERS["member_state_placement"])
            member_states_available = ident.get(HEADERS["member_states_available"])
            if not member_state_placement:
                errors.append("BR-UDID-043 / BR-UDID-045: Member State for placement on market is required for Class IIa/IIb/III devices on market.")
            if not member_states_available:
                errors.append("BR-UDID-043 / BR-UDID-673 / BR-UDID-674: Member States where device is made available are required for Class IIa/IIb/III devices on market.")

        # BR-UDID-113: Directive Certificates (conditional by risk class and legislation)
        # For now: Class III requires certificate; note on optional cases
        # In legacy MDD Class I (non-measuring) and IVDD General, certificates are optional
        if risk_class == "CLASS_III" and cert_applicable != "true":
            errors.append("BR-UDID-113: Class III devices require certificate information. Set 'Certificate applicable' to Yes.")

        # BR-UDID-705: Special Device Type cannot be set with Kit or System/SPP
        special_type = basic.get(HEADERS["special_device_type"])
        is_system_spp = self._to_bool(basic.get(HEADERS["is_system_spp"]))
        is_kit = self._to_bool(basic.get(HEADERS["is_kit"]))
        if special_type and special_type.strip():
            if is_system_spp == "true" or is_kit == "true":
                errors.append("BR-UDID-705: Special Device Type cannot be set if the device is marked as Kit or System/Procedure pack.")

        # BR-UDID-131: Additional description mandatory for System/SPP/Kit
        additional_description = ident.get(HEADERS["additional_description"])
        if is_system_spp == "true" or is_kit == "true":
            if not additional_description or not additional_description.strip():
                errors.append("BR-UDID-131: Additional product description is mandatory for System, Procedure pack, or Kit devices.")

        # BR-UDID-635: Class IIb implantable → device type specification required
        if risk_class == "CLASS_IIB" and implantable == "true":
            device_type_checkbox = self._to_bool(basic.get(HEADERS["iib_exception_flag"]))
            device_type_spec = basic.get(HEADERS["iib_exception_type"])
            if device_type_checkbox == "true":
                if not device_type_spec or not device_type_spec.strip():
                    errors.append("BR-UDID-635: For Class IIb implantable devices, the device type must be specified (suture, staple, dental filling, dental brace, tooth crown, screw, wedge, plate, wire, pin, clip, or connector).")
                # Validate it's one of the allowed types
                allowed_types = {"suture", "staple", "dental filling", "dental brace", "tooth crown", "screw", "wedge", "plate", "wire", "pin", "clip", "connector"}
                device_type_normalized = device_type_spec.strip().lower()
                if device_type_normalized not in allowed_types:
                    errors.append(f"BR-UDID-635: Device type '{device_type_spec}' is not in the allowed list. Must be one of: {', '.join(sorted(allowed_types))}.")

        if errors:
            raise ValueError("Input validation failed:\n- " + "\n- ".join(errors))
        if warnings:
            print("Validation warnings:\n- " + "\n- ".join(warnings))

    def _device_node(self, root: etree._Element) -> etree._Element:
        node = root.find(f".//{self._ns_tag_key('device.device')}")
        if node is None:
            raise ValueError("Device payload node was not created.")
        return node

    def _build_message_root(
        self,
        recipient_node_actor_code: str,
        sender_node_actor_code: str,
        conversation_id: str,
        correlation_id: str,
        message_id: str,
    ) -> etree._Element:
        # Caller must provide envelope IDs and node actor codes explicitly.
        root = etree.Element(self._ns_tag_key("message.pull_response"), nsmap=XML_NS_MAP)
        root.set("version", "3.0.28")

        self._append_text_with_default_comment(
            root,
            *self._tag_key_parts("message.conversation_id"),
            conversation_id,
            DEFAULT_CONVERSATION_ID,
            "Provide your own conversationID (default value applied)",
        )
        self._append_text_with_default_comment(
            root,
            *self._tag_key_parts("message.correlation_id"),
            correlation_id,
            DEFAULT_CORRELATION_ID,
            "Provide your own correlationID (default value applied)",
        )
        self._append_tag(root, "message.creation_datetime", datetime.now(timezone.utc).isoformat(timespec="milliseconds"))
        self._append_text_with_default_comment(
            root,
            *self._tag_key_parts("message.message_id"),
            message_id,
            DEFAULT_MESSAGE_ID,
            "Provide your own messageID (default value applied)",
        )

        recipient = self._append_tag(root, "message.recipient")
        recipient_node = self._append_tag(recipient, "message.node")
        self._append_text_with_default_comment(
            recipient_node,
            *self._tag_key_parts("service.node_actor_code"),
            recipient_node_actor_code,
            DEFAULT_RECIPIENT_NODE_ACTOR_CODE,
            "Provide your own recipient nodeActorCode (default value applied)",
        )
        recipient_service = self._append_tag(recipient, "message.service")
        self._append_tag(recipient_service, "service.service_id", XML_FIXED["service_id"])
        self._append_tag(recipient_service, "service.service_operation", XML_FIXED["service_operation"])

        payload = self._append_tag(root, "message.payload")
        device = etree.SubElement(payload, self._ns_tag_key("device.device"))
        device.set(self._ns_tag("xsi", "type"), "device:MDRDeviceType")

        sender = self._append_tag(root, "message.sender")
        sender_node = self._append_tag(sender, "message.node")
        self._append_text_with_default_comment(
            sender_node,
            *self._tag_key_parts("service.node_actor_code"),
            sender_node_actor_code,
            DEFAULT_SENDER_NODE_ACTOR_CODE,
            "Provide your own sender nodeActorCode (default value applied)",
        )
        sender_service = self._append_tag(sender, "message.service")
        self._append_tag(sender_service, "service.service_id", XML_FIXED["service_id"])
        self._append_tag(sender_service, "service.service_operation", XML_FIXED["service_operation"])

        self._append_tag(root, "message.number_of_pages", "1")
        self._append_tag(root, "message.page_number", "0")
        self._append_tag(root, "message.page_size", "20")
        self._append_tag(root, "message.response_code", XML_FIXED["response_code"])
        return root

    def _basic_udi_model_choice(self, parent: etree._Element, basic: SheetData) -> None:
        model_applicable = self._to_bool(basic.get(HEADERS["model_applicable"]))
        model_value = basic.get(HEADERS["device_model"])
        device_name = basic.get(HEADERS["device_name"])

        if model_value and device_name:
            model_name = self._append_tag(parent, "budi.model_name")
            self._append_tag(model_name, "commondevice.model", model_value)
            self._append_tag(model_name, "commondevice.name", device_name)
            return

        if model_value and model_applicable != "false":
            self._append_tag(parent, "budi.model", model_value)
            return

        if device_name:
            model_name = self._append_tag(parent, "budi.model_name")
            self._append_tag(model_name, "commondevice.name", device_name)
            return

        if model_value:
            self._append_tag(parent, "budi.model", model_value)
            return

        raise ValueError("Either 'Device Model' or 'Device Name*' must be provided.")

    def _append_basic_udi_identifier(self, parent: etree._Element, basic: SheetData, ident: SheetData) -> None:
        basic_identifier_code = basic.get("Basic UDI-DI code") or basic.get("Basic UDI code") or ident.get(HEADERS["udi_di_code"])
        issuing_entity = basic.get(HEADERS["issuing_entity"]) or ident.get(HEADERS["issuing_entity"])

        identifier = self._append_tag(parent, "budi.identifier")
        self._append_required_text_tag(identifier, "commondevice.di_code", basic_identifier_code, "Basic UDI-DI code is required.")
        self._append_required_text_tag(identifier, "commondevice.issuing_entity_code", issuing_entity, "Issuing entity is required for the Basic UDI-DI.")

    def _append_certificate_links(self, parent: etree._Element, ident: SheetData, cert: SheetData, risk_class: str) -> None:
        cert_applicable = self._to_bool(cert.get(HEADERS["cert_applicable"]))
        if risk_class == "CLASS_III" and cert_applicable != "true":
            raise ValueError("Certificate information is required for Class III devices. Set 'Certificate applicable' to Yes.")

        if cert_applicable != "true":
            return

        nb_actor_code = cert.get(HEADERS["nb_actor_code"])
        certificate_type = cert.get(HEADERS["cert_type"])
        certificate_number = cert.get(HEADERS["cert_number"]) or ident.get(HEADERS["udi_di_code"])

        if not nb_actor_code:
            raise ValueError("'Notified Body Actor Code (NBActorCode)' is required when certificate is applicable.")
        if not certificate_type:
            raise ValueError("'Certificate type' is required when certificate is applicable.")

        links_parent = self._append_tag(parent, "budi.device_certificate_links")
        link = self._append_tag(links_parent, "lnks.device_certificate_link")
        self._append_tag(link, "lnks.certificate_number", certificate_number)
        expiry_date = cert.get(HEADERS["cert_expiry"])
        if expiry_date:
            self._append_tag(link, "lnks.expiry_date", expiry_date)

        self._append_tag(link, "lnks.nb_actor_code", self._normalize_nb_actor_code(nb_actor_code))
        revision_number = cert.get(HEADERS["cert_revision"])
        if revision_number:
            self._append_tag(link, "lnks.certificate_revision_number", revision_number)

        self._append_tag(link, "lnks.certificate_type", self._normalize_certificate_type(certificate_type))

    def _build_basic_udi(self, root: etree._Element, basic: SheetData, device: SheetData, ident: SheetData, cert: SheetData) -> etree._Element:
        device_node = self._device_node(root)
        basic_udi = self._append_tag(device_node, "device.mdr_basic_udi")

        self._append_tag(basic_udi, "entity.state", XML_FIXED["state_registered"])
        self._append_tag(basic_udi, "entity.version", XML_FIXED["version"])
        self._append_tag(basic_udi, "entity.version_date", datetime.now(timezone.utc).isoformat(timespec="milliseconds"))

        risk_class = self._normalize_risk_class(basic.get(HEADERS["risk_class"]))
        if not risk_class:
            raise ValueError("'Risk class*' is required.")
        self._append_tag(basic_udi, "budi.risk_class", risk_class)

        self._basic_udi_model_choice(basic_udi, basic)
        self._append_basic_udi_identifier(basic_udi, basic, ident)

        self._append_mapped_booleans(
            basic_udi,
            {"basic": basic, "device": device},
            BASIC_UDI_BOOL_MAPPINGS[:2],
        )

        manufacturer_srn = basic.get(HEADERS["manufacturer_srn"]) or basic.get(HEADERS["mf_actor_code"])
        self._append_text_with_default_comment(
            basic_udi,
            *self._tag_key_parts("budi.mf_actor_code"),
            manufacturer_srn,
            DEFAULT_MF_ACTOR_CODE,
            "Provide your own MFActorCode (Manufacturer Actor Code / SRN) (default value applied)",
        )

        self._append_certificate_links(basic_udi, ident, cert, risk_class)

        self._append_mapped_booleans(
            basic_udi,
            {"basic": basic, "device": device},
            BASIC_UDI_BOOL_MAPPINGS[2:3],
        )
        iib_exception_flag = basic.get(HEADERS["iib_exception_flag"])
        self._append_bool_tag(basic_udi, "budi.iib_implantable_exceptions", iib_exception_flag, default="false")
        iib_exception_type = basic.get(HEADERS["iib_exception_type"])
        if self._to_bool(iib_exception_flag) == "true" and iib_exception_type:
            # The XSD only exposes a boolean IIb exception field; keep the selected type as XML comment.
            basic_udi.append(etree.Comment(f"IIb exception type selected by user: {iib_exception_type}"))
        self._append_mapped_booleans(
            basic_udi,
            {"basic": basic, "device": device},
            BASIC_UDI_BOOL_MAPPINGS[3:4],
        )

        self._append_tag(basic_udi, "budi.type", XML_FIXED["device_type"])

        # Apply BR-UDID auto-corrections at write time.
        implantable_value = self._to_bool(basic.get(HEADERS["implantable"])) or "false"
        if risk_class == "CLASS_I" and implantable_value == "true":
            implantable_value = "false"

        reusable_value = self._to_bool(basic.get(HEADERS["reusable"])) or "false"
        if implantable_value == "true" and reusable_value == "true":
            reusable_value = "false"

        self._append_mapped_booleans(
            basic_udi,
            {"basic": basic, "device": device},
            BASIC_UDI_BOOL_MAPPINGS[4:6],
        )
        self._append_tag(basic_udi, "commondevice.implantable", implantable_value)
        self._append_mapped_booleans(
            basic_udi,
            {"basic": basic, "device": device},
            BASIC_UDI_BOOL_MAPPINGS[7:8],
        )
        self._append_tag(basic_udi, "commondevice.reusable", reusable_value)

        return basic_udi

    def _build_udidi_data(self, root: etree._Element, ident: SheetData, chars: SheetData, device: SheetData, basic: SheetData) -> etree._Element:
        device_node = self._device_node(root)
        udidi_data = self._append_tag(device_node, "device.mdr_udidi_data")

        self._append_tag(udidi_data, "entity.state", XML_FIXED["state_registered"])
        self._append_tag(udidi_data, "entity.version", XML_FIXED["version"])
        self._append_tag(udidi_data, "entity.version_date", datetime.now(timezone.utc).isoformat(timespec="milliseconds"))

        identifier = self._append_tag(udidi_data, "udidi.identifier")
        self._append_required_text_tag(identifier, "commondevice.di_code", ident.get(HEADERS["udi_di_code"]), "UDI-DI code is required.")
        self._append_required_text_tag(identifier, "commondevice.issuing_entity_code", ident.get(HEADERS["issuing_entity"]), "UDI-DI issuing entity is required.")

        status = self._append_tag(udidi_data, "udidi.status")
        self._append_text_with_default_comment(
            status,
            *self._tag_key_parts("commondevice.code"),
            self._normalize_status(ident.get(HEADERS["udi_di_status"])),
            DEFAULT_UDI_STATUS_CODE,
            "Provide your own UDI-DI status code (default value applied)",
        )

        basic_identifier = self._append_tag(udidi_data, "udidi.basic_udi_identifier")
        basic_code = basic.get("Basic UDI-DI code") or basic.get("Basic UDI code") or ident.get(HEADERS["udi_di_code"])
        basic_issuer = basic.get(HEADERS["issuing_entity"]) or ident.get(HEADERS["issuing_entity"])
        self._append_required_text_tag(basic_identifier, "commondevice.di_code", basic_code, "Basic UDI-DI code is required.")
        self._append_required_text_tag(basic_identifier, "commondevice.issuing_entity_code", basic_issuer, "Basic UDI-DI issuing entity is required.")

        self._append_text_with_default_comment(
            udidi_data,
            *self._tag_key_parts("udidi.mdn_codes"),
            ident.get(HEADERS["emdn_code"]),
            DEFAULT_MDN_CODE,
            "Provide your own EMDN code for MDNCodes (default value applied)",
        )

        production_tokens: list[str] = []
        if self._to_bool(ident.get(HEADERS["pi_lot"])) == "true":
            production_tokens.append("BATCH_NUMBER")
        if self._to_bool(ident.get(HEADERS["pi_serial"])) == "true":
            production_tokens.append("SERIALISATION_NUMBER")
        if self._to_bool(ident.get(HEADERS["pi_mfg_date"])) == "true":
            production_tokens.append("MANUFACTURING_DATE")
        if self._to_bool(ident.get(HEADERS["pi_exp_date"])) == "true":
            production_tokens.append("EXPIRATION_DATE")
        if production_tokens:
            self._append_tag(udidi_data, "udidi.production_identifier", " ".join(production_tokens))

        self._append_text_with_default_comment(
            udidi_data,
            *self._tag_key_parts("udidi.reference_number"),
            ident.get(HEADERS["reference_number"]) or ident.get(HEADERS["udi_di_code"]),
            DEFAULT_REFERENCE_NUMBER,
            "Provide your own reference/catalogue number (default value applied)",
        )

        secondary_applicable = self._to_bool(ident.get(HEADERS["secondary_applicable"]))
        sec_issuer = ident.get(HEADERS["secondary_issuer"])
        sec_code = ident.get(HEADERS["secondary_code"])
        if secondary_applicable == "true" and sec_issuer and sec_code:
            secondary = self._append_tag(udidi_data, "udidi.secondary_identifier")
            self._append_required_text_tag(secondary, "commondevice.di_code", sec_code, "Secondary UDI-DI code is required.")
            self._append_required_text_tag(secondary, "commondevice.issuing_entity_code", sec_issuer, "Secondary UDI-DI issuer is required.")

        self._append_mapped_booleans(
            udidi_data,
            {"chars": chars, "device": device},
            UDIDI_BOOL_MAPPINGS[:2],
        )

        trade_name = ident.get(HEADERS["trade_name"])
        if trade_name:
            trade_names = self._append_tag(udidi_data, "udidi.trade_names")
            trade_name_entry = self._append_tag(trade_names, "lngs.name")
            self._append_tag(trade_name_entry, "lngs.language", self._normalize_language(ident.get(HEADERS["language"]), default="ANY"))
            self._append_tag(trade_name_entry, "lngs.text_value", trade_name)

        additional_description = ident.get(HEADERS["additional_description"])
        if additional_description:
            additional = self._append_tag(udidi_data, "udidi.additional_description")
            additional_entry = self._append_tag(additional, "lngs.name")
            self._append_tag(additional_entry, "lngs.language", self._normalize_language(ident.get(HEADERS["language"], occurrence=2), default="ANY"))
            self._append_tag(additional_entry, "lngs.text_value", additional_description)

        website = ident.get(HEADERS["website"])
        if website:
            self._append_tag(udidi_data, "udidi.website", website)

        single_use = self._to_bool(chars.get(HEADERS["single_use"]))
        max_reuses = chars.get(HEADERS["max_reuses"])
        if single_use == "true":
            self._append_tag(udidi_data, "udidi.number_of_reuses", "0")
        elif max_reuses and str(max_reuses).isdigit():
            self._append_tag(udidi_data, "udidi.number_of_reuses", str(max_reuses))
        else:
            self._append_tag(udidi_data, "udidi.number_of_reuses", "1")

        self._append_tag(udidi_data, "udidi.base_quantity", ident.get(HEADERS["quantity"]) or "1")
        self._append_mapped_booleans(
            udidi_data,
            {"chars": chars, "device": device},
            UDIDI_BOOL_MAPPINGS[2:],
        )

        if self._to_bool(chars.get(HEADERS["clinical_size_status"])) == "true":
            pass

        return udidi_data

    def build_xml_from_excel(
        self,
        excel_file: Path,
        xsd_file: Path,
        recipient_node_actor_code: str,
        sender_node_actor_code: str,
        conversation_id: str,
        correlation_id: str,
        message_id: str,
    ) -> etree._ElementTree:
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        basic_ws = self._sheet_by_prefix(wb, SHEET_PREFIXES["basic"])
        cert_ws = self._optional_sheet_by_prefix(wb, SHEET_PREFIXES["cert"])
        ident_ws = self._sheet_by_prefix(wb, SHEET_PREFIXES["ident"])
        chars_ws = self._sheet_by_prefix(wb, SHEET_PREFIXES["chars"])
        device_ws = self._sheet_by_prefix(wb, SHEET_PREFIXES["device"])

        basic = self._sheet_data(basic_ws)
        cert = self._sheet_data(cert_ws) if cert_ws is not None else SheetData(headers=[], values=[])
        ident = self._sheet_data(ident_ws)
        chars = self._sheet_data(chars_ws)
        device = self._sheet_data(device_ws)

        constraints = self._load_schema_constraints(xsd_file)
        self._validate_input_data(basic, ident, chars, device, cert, constraints)

        recipient_code = (recipient_node_actor_code or "").strip()
        sender_code = (sender_node_actor_code or "").strip()

        root = self._build_message_root(
            recipient_node_actor_code=recipient_code,
            sender_node_actor_code=sender_code,
            conversation_id=conversation_id,
            correlation_id=correlation_id,
            message_id=message_id,
        )
        self._build_basic_udi(root, basic, device, ident, cert)
        self._build_udidi_data(root, ident, chars, device, basic)
        return etree.ElementTree(root)

    def _validate_xsd(self, xml_file: Path, xsd_file: Path) -> None:
        if not xsd_file.exists():
            raise FileNotFoundError(f"XSD not found: {xsd_file}")

        original_cwd = os.getcwd()
        xsd_root = xsd_file.parent.parent.parent
        try:
            os.chdir(xsd_root)
            xsd_doc = etree.parse(str(xsd_file.relative_to(xsd_root)))
            schema = etree.XMLSchema(xsd_doc)
            xml_doc = etree.parse(str(xml_file))
            if not schema.validate(xml_doc):
                issues = [f"Line {error.line}, Col {error.column}: {error.message}" for error in schema.error_log]
                raise ValueError("XSD validation failed:\n" + "\n".join(issues))
            print("XSD validation: PASS")
        finally:
            os.chdir(original_cwd)

    def generate(
        self,
        excel_path: str,
        output_xml_path: str,
        xsd_path: str,
        recipient_node_actor_code: str,
        sender_node_actor_code: str,
        conversation_id: str,
        correlation_id: str,
        message_id: str,
    ) -> Path:
        excel_file = self._resolve(excel_path)
        output_file = self._resolve(output_xml_path)
        xsd_file = self._resolve(xsd_path)

        if not excel_file.exists():
            raise FileNotFoundError(f"Excel not found: {excel_file}")

        tree = self.build_xml_from_excel(
            excel_file,
            xsd_file,
            recipient_node_actor_code=recipient_node_actor_code,
            sender_node_actor_code=sender_node_actor_code,
            conversation_id=conversation_id,
            correlation_id=correlation_id,
            message_id=message_id,
        )
        output_file.parent.mkdir(parents=True, exist_ok=True)
        tree.write(str(output_file), pretty_print=True, xml_declaration=True, encoding="utf-8")
        self._validate_xsd(output_file, xsd_file)
        return output_file


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Generate EUDAMED XML from Excel and validate against XSD.")
    parser.add_argument("--excel", required=True, help="Input Excel path")
    parser.add_argument("--output", required=True, help="Output XML path")
    parser.add_argument(
        "--xsd",
        default="XSD_schemas/service/Message.xsd",
        help="Root XSD for validation (default: XSD_schemas/service/Message.xsd)",
    )
    parser.add_argument(
        "--recipient-node-actor-code",
        default="",
        help="message.recipient.node.nodeActorCode. Leave empty to emit blank value with XML comment.",
    )
    parser.add_argument(
        "--sender-node-actor-code",
        default="",
        help="message.sender.node.nodeActorCode. Leave empty to emit blank value with XML comment.",
    )
    parser.add_argument(
        "--conversation-id",
        default="",
        help="message.conversationID. Leave empty to emit blank value with XML comment.",
    )
    parser.add_argument(
        "--correlation-id",
        default="",
        help="message.correlationID. Leave empty to emit blank value with XML comment.",
    )
    parser.add_argument(
        "--message-id",
        default="",
        help="message.messageID. Leave empty to emit blank value with XML comment.",
    )
    return parser


def main() -> None:
    args = build_parser().parse_args()
    workspace_root = Path(__file__).resolve().parent
    generator = ExcelToEudamedXML(workspace_root)
    output = generator.generate(
        excel_path=args.excel,
        output_xml_path=args.output,
        xsd_path=args.xsd,
        recipient_node_actor_code=args.recipient_node_actor_code,
        sender_node_actor_code=args.sender_node_actor_code,
        conversation_id=args.conversation_id,
        correlation_id=args.correlation_id,
        message_id=args.message_id,
    )
    print(f"Generated XML: {output}")


if __name__ == "__main__":
    main()
