#!/usr/bin/env python3
"""Create a user-fillable Excel template for EUDAMED XML generation."""

from __future__ import annotations

import argparse
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill


HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)


SHEETS: dict[str, list[str]] = {
    "Basic UDI-DI information": [
        "Risk class*",
        "Device model applicable",
        "Device Model",
        "Device Name*",
        "Implantable*",
        "Measuring function*",
        "Reusable surgical instrument*",
        "Active device*",
        "Device intended to administer and/or remove medicinal product*",
        "Is it a System or Procedure pack which is a Device in itself (Y/N)",
        "Is it a Kit (Y/N)",
        "Special Device Type (if applicable)",
        "IIb implantable exceptions: Is device a suture/staple/dental/screw/etc (Y/N)",
        "IIb implantable exceptions: Specify device type (suture, staple, dental filling, screw, etc.)",
    ],
    "Certificate information": [
        "Certificate applicable",
        "Notified Body Actor Code (NBActorCode)",
        "Certificate type",
        "Certificate number",
        "Certificate revision number",
        "Certificate expiry date (YYYY-MM-DD)",
    ],
    "UDI-DI identification": [
        "UDI-DI identification Issuing Entity (GS1/HIBCC/ICCBBA/IFA)",
        "UDI-DI code",
        "Reference/Catalogue number",
        "URL for additional information (as electronic instructions for use)",
        "UDI-DI status ( On the EU market / No longer placed on the EU market / Not intended for the EU market )",
        "Member State where device is placed on market (e.g., IT, DE, FR)",
        "Member States where device is made available (comma-separated, e.g., IT,DE,FR)",
        "Quantity of device",
        "UDI-DI from another entity (secondary) applicable",
        "Secondary Issuing Entity (GS1/HIBCC/ICCBBA/IFA)",
        "Secondary UDI-DI code",
        "Enter a nomenclature code (EMDN code)",
        "Trade name",
        "Select the language",
        "Additional product description",
        "Select the language",
        "Type of UDI-PI Lot or Batch number",
        "Type of UDI-PI Serial number",
        "Type of UDI-PI Manufacturing date",
        "Type of UDI-PI Expiration date",
    ],
    "UDI-DI characteristics": [
        "Need for sterilisation before use",
        "Device labelled as sterile",
        "Containing latex",
        "Labelled as single use",
        "Maximum number of reuses",
        "Clinical size application status",
    ],
    "Device information": [
        "Tissues and cells Presence of human tissues or cells, or their derivatives:",
        "Tissues and cells Presence of animal tissues or cells, or their derivatives:",
        "Information on substances Presence of a substance which, if used separately, may be considered to be a medicinal product (Y/N)",
        "Presence of a substance which, if used separately, may be considered to be a medicinal product derived from human blood or human plasma(Y/N)",
        "Reprocessed single use device",
    ],
}


EXAMPLES: dict[str, dict[str, str]] = {
    "Basic UDI-DI information": {
        "Risk class*": "Class IIa",
        "Device model applicable": "Yes",
        "Device Model": "MODEL-100",
        "Device Name*": "Example Device",
        "Implantable*": "No",
        "Measuring function*": "Yes",
        "Reusable surgical instrument*": "No",
        "Active device*": "No",
        "Device intended to administer and/or remove medicinal product*": "No",
        "Is it a System or Procedure pack which is a Device in itself (Y/N)": "No",
        "Is it a Kit (Y/N)": "No",
        "Special Device Type (if applicable)": "",
        "IIb implantable exceptions: Is device a suture/staple/dental/screw/etc (Y/N)": "No",
        "IIb implantable exceptions: Specify device type (suture, staple, dental filling, screw, etc.)": "",
    },
    "UDI-DI identification": {
        "UDI-DI identification Issuing Entity (GS1/HIBCC/ICCBBA/IFA)": "GS1",
        "UDI-DI code": "01234567890123",
        "UDI-DI status ( On the EU market / No longer placed on the EU market / Not intended for the EU market )": "On the EU market",
        "Member State where device is placed on market (e.g., IT, DE, FR)": "IT",
        "Member States where device is made available (comma-separated, e.g., IT,DE,FR)": "IT,DE,FR",
        "Enter a nomenclature code (EMDN code)": "A0000",
        "Trade name": "Example Trade Name",
        "Select the language": "EN",
        "Type of UDI-PI Lot or Batch number": "Yes",
    },
    "UDI-DI characteristics": {
        "Need for sterilisation before use": "No",
        "Device labelled as sterile": "No",
        "Containing latex": "No",
        "Labelled as single use": "Yes",
        "Maximum number of reuses": "0",
        "Clinical size application status": "No",
    },
    "Certificate information": {
        "Certificate applicable": "Yes",
        "Notified Body Actor Code (NBActorCode)": "NB 0123",
        "Certificate type": "MDR_TYPE_EXAMINATION",
        "Certificate number": "CERT-2026-0001",
        "Certificate revision number": "1",
        "Certificate expiry date (YYYY-MM-DD)": "2030-12-31",
    },
}


GUIDANCE_ROWS = [
    (
        "Risk class*",
        "All",
        "Use values such as Class I, Class IIa, Class IIb, Class III. (BR-UDID-676: Class I forces Implantable=No)",
    ),
    (
        "Device Name* or Device Model",
        "All",
        "At least one should be provided. (BR-UDID-066)",
    ),
    (
        "Implantable*",
        "All",
        "If Yes and Risk Class IIb, device type must be specified (BR-UDID-635). If Yes, Reusable surgical instrument auto-forced to No (BR-UDID-677).",
    ),
    (
        "IIb implantable exceptions",
        "Class IIb with Implantable=Yes",
        "Required for Class IIb implantable devices (BR-UDID-635). Specify: suture, staple, dental filling, dental brace, tooth crown, screw, wedge, plate, wire, pin, clip, or connector.",
    ),
    (
        "Labelled as single use",
        "All",
        "If No, then Max reuses must be provided. (BR-UDID-024)",
    ),
    (
        "Maximum number of reuses",
        "If single use = No",
        "Only applicable if NOT single-use. (BR-UDID-024)",
    ),
    (
        "Member State fields",
        "Class IIa, IIb, III (when on market)",
        "Required for placed on market / made available. (BR-UDID-043, BR-UDID-673, BR-UDID-674)",
    ),
    (
        "Is it a System/SPP/Kit",
        "Conditional",
        "If System/SPP=Yes or Kit=Yes, then Special Device Type cannot be set. (BR-UDID-705)",
    ),
    (
        "Additional product description",
        "If System/SPP/Kit",
        "Mandatory for System, Procedure pack, or Kit devices. (BR-UDID-131)",
    ),
    (
        "UDI-DI code + Issuing Entity",
        "All",
        "Must uniquely identify the UDI-DI. (BR-UDID-003)",
    ),
    (
        "UDI-DI status",
        "All",
        "Use one of: On the EU market, No longer placed on the EU market, Not intended for the EU market. (BR-UDID-458, BR-UDID-073)",
    ),
    (
        "Trade name + language",
        "All",
        "Language should be 2/3-letter code (for example EN, FR, ANY). (BR-UDID-070)",
    ),
    (
        "Certificate information",
        "Class III (required), Class I with measuring (required), others (optional)",
        "If Certificate applicable is Yes, provide NBActorCode and Certificate type at minimum. (BR-UDID-113)",
    ),
]


def build_workbook() -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for sheet_name, headers in SHEETS.items():
        ws = wb.create_sheet(sheet_name)
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = max(18, min(72, len(header) + 2))

            sample_value = EXAMPLES.get(sheet_name, {}).get(header, "")
            ws.cell(row=2, column=col, value=sample_value)

    guide = wb.create_sheet("Guidance")
    guide_headers = ["Field", "Required For", "Notes"]
    for col, header in enumerate(guide_headers, start=1):
        cell = guide.cell(row=1, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        guide.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 40

    for row, values in enumerate(GUIDANCE_ROWS, start=2):
        for col, value in enumerate(values, start=1):
            guide.cell(row=row, column=col, value=value)

    return wb


def main() -> None:
    parser = argparse.ArgumentParser(description="Create a fillable Excel template for EUDAMED XML generation.")
    parser.add_argument("--output", required=True, help="Output .xlsx path")
    args = parser.parse_args()

    output = Path(args.output)
    if not output.is_absolute():
        output = Path(__file__).resolve().parent / output

    output.parent.mkdir(parents=True, exist_ok=True)
    wb = build_workbook()
    wb.save(output)
    print(f"Template created: {output}")


if __name__ == "__main__":
    main()
